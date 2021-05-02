from django.shortcuts import render
from .utils import get_plot
import csv
import openpyxl
def read():
    data=[]
    wb = openpyxl.load_workbook(r'C:\Users\sinch\AppData\Local\Programs\Python\Python39\Scripts\test_web\test_app\state_wise.xlsx')
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        data.append(row)
    return data
# Create your views here.
def offers(request):
    '''with open('state_wise.csv', 'r') as f:
        reader = csv.reader(f)
        data = list(reader)'''
    
    x=[]
    y=[]
    d={}
    for i in range(2,7):
        x.append(data[i][0])
        y.append(data[i][4])
        d[data[i][0]]=data[i][4]
    return render(request,'offers_real.html',{"context": d})


'''class ChartData(APIView):
    authentication_classes = []
    permission_classes = []

    def get(self, request, format=None):
        qs=Company.objects.all()
        
        labels = []
        default_items = []

        for item in qs:
            labels.append(item.name)
            default_items.append(item.profit)

        data = {
                "labels": labels,
                "default": default_items,
        }
        return Response(data)'''
def homepage(request):
    data=read()
    total=data[1][1]
    recoveries=data[1][2]
    deaths=data[1][3]
    active=data[1][4]
    total_percent=str(total/14000000)
    recoveries_percent=str(recoveries/active)
    active_percent=str(active/total)
    deaths_percent=str(deaths/active)
    recoveries_percent=recoveries_percent[:4]
    active_percent=active_percent[:4]
    deaths_percent=deaths_percent[:4]
    total_percent=total_percent[:4]
    d={}
    d['total']=total
    d['recoveries']=recoveries
    d['deaths']=deaths
    d['active']=active
    d['total_percent']=total_percent
    d['recoveries_percent']=recoveries_percent
    d['active_percent']=active_percent
    d['deaths_percent']=deaths_percent
    data2=[]
    wb = openpyxl.load_workbook(r'C:\Users\sinch\AppData\Local\Programs\Python\Python39\Scripts\test_web\test_app\beds.xlsx')
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        data2.append(row)
    d['bedtotal']=data2[38][3]
    data3=[]
    wb = openpyxl.load_workbook(r'C:\Users\sinch\AppData\Local\Programs\Python\Python39\Scripts\test_web\test_app\vaccine.xlsx')
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        data3.append(row)
    d['vaxtotal']=data3[38][103]
    d['percent']=str(data3[38][103]/14000000)
    d['percent']=d['percent'][:4]
    p=[]
    for i in range(2,7):
        d2={}
        state=data[i][0]
        tot=data[i][1]
        act=data[i][4]
        date=data[i][5]
        covax=[i for i in data3 if i[0]==state]
        q=[j for j in data2 if j[0]==state]
        vac=covax[-1][-1]
        bed=q[-1][-1]
        d2['state']=state
        d2['tot']=tot
        d2['act']=act
        d2['date']=date
        d2['vac']=vac
        d2['bed']=bed
        ratio=act/bed
        if ratio<2:
            d2['values']=0
        else:
            d2['values']=2
        p.append(d2)
        
    return render(request,'index.html',{'india':d,'p':p})






